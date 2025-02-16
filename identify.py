import cognitive_face as CF
from global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
import time


#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")

sheet = wb.get_sheet_by_name('Ece 19 Final Year')

def getDateColumn():
	for i in range(1, len(sheet[1]) + 1):
		col = get_column_letter(i)
		if sheet['%s%s'% (col,'1')].value == currentDate:
			return col
			
			
Key = '6b2222d0eb264de99fb64b95ef49b851'
CF.Key.set(Key)
BASE_URL = 'https://centralindia.api.cognitive.microsoft.com/face/v1.0'
CF.BaseUrl.set(BASE_URL)

connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(80)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
		res = CF.face.detect(imgurl)
		if len(res) != 1:
			print ("No face detected.")
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		#print (filename)
		#print (res)
		for face  in res:
			if not face['candidates']:
				print ("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
				row = c.fetchone()
				if row is not None:
					attend[int(row[0])] += 1
					print( row[1] + " recognized")
		time.sleep(6)
		
count=0
for row in range(2, len(sheet['A']) + 1):
	rn = sheet['A%s'% row].value
	if rn is not None:
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1
			count=count+1
print("No. of presentees",count)
print("Total no. of students",len(sheet['A'])-2)
wb.save(filename = "reports.xlsx")	 	
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
